"""
Job Scraper — aryan2k1
Sources: Adzuna + Reed APIs, plus Greenhouse/Lever public APIs (no auth)
Run via GitHub Actions cron or locally: python scraper.py
"""

import os
import time
import logging
from datetime import date, datetime
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────

def _load_dotenv(path: str = ".env") -> None:
    """Load KEY=VALUE lines from a local .env file (no external dependency).
    In GitHub Actions the values come from repository secrets instead."""
    if not os.path.exists(path):
        return
    for line in open(path, encoding="utf-8"):
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, _, value = line.partition("=")
            os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


_load_dotenv()

ADZUNA_APP_ID  = os.environ.get("ADZUNA_APP_ID", "")
ADZUNA_APP_KEY = os.environ.get("ADZUNA_APP_KEY", "")
REED_API_KEY   = os.environ.get("REED_API_KEY", "")
OUTPUT_FILE    = os.environ.get("OUTPUT_FILE", "job_tracker.xlsx")

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; JobScraper/1.0)"}

# Keywords that must appear in job title
INCLUDE_KEYWORDS = [
    "analyst", "associate", "graduate", "credit", "ratings",
    "analytics", "data", "client", "investor", "consultant",
    "research", "finance", "risk", "quantitative",
]

# Titles to skip — too senior or wrong fit
EXCLUDE_KEYWORDS = [
    "senior", "director", "head of", "vp ", "vice president",
    "principal", "managing director", "partner", "md ",
    "executive director", "c-suite", "cfo", "ceo", "coo",
]

# ── Target companies ──────────────────────────────────────────────────────────
# (display_name, tier, adzuna_search_name, greenhouse_slug, lever_slug)
# greenhouse_slug / lever_slug = None if they don't use that ATS

TARGETS = [
    # Tier 1 — Ratings
    ("Fitch Ratings",       1, "Fitch Ratings",    None,           None),
    ("S&P Global",          1, "S&P Global",        "spglobal",     None),
    ("DBRS Morningstar",    1, "Morningstar",       "morningstar",  None),
    ("Kroll",               1, "Kroll",             "kroll",        None),

    # Tier 1 — Financial data
    ("MSCI",                1, "MSCI",              "msci",         None),
    ("FactSet",             1, "FactSet",           "factset",      None),
    ("LSEG",                1, "LSEG",              None,           None),
    ("ICE Data Services",   1, "ICE",               None,           None),

    # Tier 1 — Asset managers
    ("Schroders",           1, "Schroders",         None,           None),
    ("Abrdn",               1, "Abrdn",             None,           None),
    ("LGIM",                1, "Legal and General", None,           None),
    ("Ninety One",          1, "Ninety One",        "ninetyone",    None),

    # Tier 2 — Big 4
    ("Deloitte",            2, "Deloitte",          None,           None),
    ("EY",                  2, "Ernst Young",       None,           None),
    ("KPMG",                2, "KPMG",              None,           None),
    ("PwC",                 2, "PricewaterhouseCoopers", None,      None),

    # Tier 2 — Other
    ("Quantexa",            2, "Quantexa",          "quantexa",     None),
    ("Barclays",            2, "Barclays",          None,           None),
    ("NatWest",             2, "NatWest",           None,           None),
    ("HSBC",                2, "HSBC",              None,           None),
    ("Revolut",             2, "Revolut",           None,           "revolut"),
]


# ── Filters ───────────────────────────────────────────────────────────────────

def is_relevant(title: str) -> bool:
    t = title.lower()
    return (
        any(k in t for k in INCLUDE_KEYWORDS) and
        not any(k in t for k in EXCLUDE_KEYWORDS)
    )


# ── Source 1: Adzuna API ──────────────────────────────────────────────────────
# Free tier: 250 calls/day. Sign up at https://developer.adzuna.com/

def adzuna_search(company_display: str, adzuna_name: str, tier: int) -> list[dict]:
    if not ADZUNA_APP_ID or not ADZUNA_APP_KEY:
        log.warning("Adzuna API credentials missing — set ADZUNA_APP_ID and ADZUNA_APP_KEY")
        return []

    jobs = []
    try:
        url = "https://api.adzuna.com/v1/api/jobs/gb/search/1"
        params = {
            "app_id":           ADZUNA_APP_ID,
            "app_key":          ADZUNA_APP_KEY,
            "company":          adzuna_name,
            "where":            "London",
            "results_per_page": 15,
            "sort_by":          "date",
            "content-type":     "application/json",
        }
        r = requests.get(url, params=params, headers=HEADERS, timeout=15)
        r.raise_for_status()
        data = r.json()

        for result in data.get("results", []):
            title = result.get("title", "")
            if not is_relevant(title):
                continue
            jobs.append({
                "company":    company_display,
                "tier":       tier,
                "title":      title,
                "url":        result.get("redirect_url", ""),
                "location":   result.get("location", {}).get("display_name", "London"),
                "found_date": str(date.today()),
                "source":     "Adzuna",
                "status":     "New",
                "applied":    "",
                "notes":      "",
            })

        log.info(f"  Adzuna  → {company_display}: {len(jobs)} match(es)")

    except Exception as e:
        # Adzuna puts the app_key in the URL, so never log the raw exception
        # (it can echo the query string with the key). Log the type only.
        log.warning(f"  Adzuna  → {company_display}: request failed ({type(e).__name__})")

    return jobs


# ── Source 2: Reed API ────────────────────────────────────────────────────────
# Free tier: unlimited search (key required). Sign up at https://www.reed.co.uk/developers/

def reed_search(company_display: str, adzuna_name: str, tier: int) -> list[dict]:
    if not REED_API_KEY:
        return []

    jobs = []
    try:
        url = "https://www.reed.co.uk/api/1.0/search"
        params = {
            "employerName": adzuna_name,
            "locationName": "London",
            "distancefromlocation": 10,
            "resultsToTake": 10,
        }
        r = requests.get(url, params=params, auth=(REED_API_KEY, ""), timeout=15)
        r.raise_for_status()
        data = r.json()

        for result in data.get("results", []):
            title = result.get("jobTitle", "")
            if not is_relevant(title):
                continue
            jobs.append({
                "company":    company_display,
                "tier":       tier,
                "title":      title,
                "url":        result.get("jobUrl", ""),
                "location":   result.get("locationName", "London"),
                "found_date": str(date.today()),
                "source":     "Reed",
                "status":     "New",
                "applied":    "",
                "notes":      "",
            })

        log.info(f"  Reed    → {company_display}: {len(jobs)} match(es)")

    except Exception as e:
        log.warning(f"  Reed    → {company_display}: {e}")

    return jobs


# ── Source 3: Greenhouse public API ──────────────────────────────────────────
# No auth needed. Works for any company using Greenhouse ATS.

def greenhouse_search(company_display: str, slug: str, tier: int) -> list[dict]:
    if not slug:
        return []

    jobs = []
    try:
        url = f"https://boards-api.greenhouse.io/v1/boards/{slug}/jobs"
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        data = r.json()

        for result in data.get("jobs", []):
            title = result.get("title", "")
            location = result.get("location", {}).get("name", "")
            if "london" not in location.lower() and "united kingdom" not in location.lower() and "uk" not in location.lower():
                continue
            if not is_relevant(title):
                continue
            jobs.append({
                "company":    company_display,
                "tier":       tier,
                "title":      title,
                "url":        result.get("absolute_url", ""),
                "location":   location,
                "found_date": str(date.today()),
                "source":     "Greenhouse",
                "status":     "New",
                "applied":    "",
                "notes":      "",
            })

        log.info(f"  GH      → {company_display}: {len(jobs)} match(es)")

    except Exception as e:
        log.warning(f"  GH      → {company_display}: {e}")

    return jobs


# ── Source 4: Lever public API ────────────────────────────────────────────────
# No auth needed. Works for any company using Lever ATS.

def lever_search(company_display: str, slug: str, tier: int) -> list[dict]:
    if not slug:
        return []

    jobs = []
    try:
        url = f"https://api.lever.co/v0/postings/{slug}?mode=json&location=London"
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        data = r.json()

        for result in data:
            title = result.get("text", "")
            location = result.get("categories", {}).get("location", "")
            if not is_relevant(title):
                continue
            jobs.append({
                "company":    company_display,
                "tier":       tier,
                "title":      title,
                "url":        result.get("hostedUrl", ""),
                "location":   location,
                "found_date": str(date.today()),
                "source":     "Lever",
                "status":     "New",
                "applied":    "",
                "notes":      "",
            })

        log.info(f"  Lever   → {company_display}: {len(jobs)} match(es)")

    except Exception as e:
        log.warning(f"  Lever   → {company_display}: {e}")

    return jobs


# ── Orchestrator ──────────────────────────────────────────────────────────────

def scrape_all() -> list[dict]:
    all_jobs = []
    seen_urls = set()

    for name, tier, adzuna_name, gh_slug, lever_slug in TARGETS:
        log.info(f"\nChecking {name}…")

        # Gather from all available sources
        candidates = (
            adzuna_search(name, adzuna_name, tier) +
            reed_search(name, adzuna_name, tier) +
            greenhouse_search(name, gh_slug, tier) +
            lever_search(name, lever_slug, tier)
        )

        # Deduplicate by URL
        for job in candidates:
            url = job.get("url", "")
            if url and url not in seen_urls:
                seen_urls.add(url)
                all_jobs.append(job)

        time.sleep(0.5)

    return all_jobs


# ── Excel ─────────────────────────────────────────────────────────────────────

NAVY  = "1F3864"
WHITE = "FFFFFF"
LGREY = "F9FAFB"

COLS = [
    ("Company",     18),
    ("Tier",         5),
    ("Role / Title", 36),
    ("URL",          40),
    ("Location",     16),
    ("Found",        11),
    ("Source",       12),
    ("Status",       13),
    ("Applied?",     10),
    ("Notes",        28),
]

STATUS_FILL = {
    "New":       "DBEAFE",
    "Applied":   "D1FAE5",
    "Interview": "FEF3C7",
    "Rejected":  "FEE2E2",
    "Offer":     "BBF7D0",
}

def _border():
    s = Side(style="thin", color="E5E7EB")
    return Border(left=s, right=s, top=s, bottom=s)


def build_excel(new_jobs: list[dict], path: str):
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        existing = {str(r[3]) for r in ws.iter_rows(min_row=2, values_only=True) if r[3]}
        to_add = [j for j in new_jobs if j.get("url") not in existing]
        log.info(f"\n{len(to_add)} new posting(s) (skipped {len(new_jobs)-len(to_add)} duplicate(s))")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Job Tracker"
        _write_header(ws)
        to_add = new_jobs
        log.info(f"\n{len(to_add)} posting(s) written")

    for job in to_add:
        _write_row(ws, job)

    for i, (_, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    wb.save(path)
    log.info(f"Saved → {path}")


def _write_header(ws):
    hf = Font(name="Calibri", bold=True, color=WHITE, size=11)
    hfill = PatternFill("solid", fgColor=NAVY)
    ha = Alignment(horizontal="center", vertical="center")
    for col, (name, _) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = _border()
    ws.row_dimensions[1].height = 22


def _write_row(ws, job: dict):
    row = ws.max_row + 1
    bg = LGREY if row % 2 == 0 else WHITE
    rfill = PatternFill("solid", fgColor=bg)
    bf = Font(name="Calibri", size=10)
    uf = Font(name="Calibri", size=10, color="1D4ED8", underline="single")

    values = [
        job.get("company", ""),
        job.get("tier", ""),
        job.get("title", ""),
        job.get("url", ""),
        job.get("location", ""),
        job.get("found_date", str(date.today())),
        job.get("source", ""),
        job.get("status", "New"),
        job.get("applied", ""),
        job.get("notes", ""),
    ]

    status_hex = STATUS_FILL.get(job.get("status", "New"), "DBEAFE")

    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = _border()
        c.alignment = Alignment(vertical="center", wrap_text=(col in (3, 4, 10)))

        if col == 4:   # URL
            c.font = uf
        elif col == 2:  # Tier
            c.fill = PatternFill("solid", fgColor="DBEAFE" if val == 1 else "D1FAE5")
            c.font = Font(name="Calibri", size=10, bold=True,
                          color="1D4ED8" if val == 1 else "065F46")
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 8:  # Status
            c.fill = PatternFill("solid", fgColor=status_hex)
            c.font = Font(name="Calibri", size=10, bold=True)
        else:
            c.font = bf
            c.fill = rfill

    ws.row_dimensions[row].height = 18


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    log.info("=" * 55)
    log.info(f"Job scraper — {datetime.now().strftime('%d %b %Y %H:%M')}")
    log.info("=" * 55)

    if not ADZUNA_APP_ID and not REED_API_KEY:
        log.error(
            "\nNo API credentials found."
            "\nSet ADZUNA_APP_ID + ADZUNA_APP_KEY (https://developer.adzuna.com)"
            "\nor REED_API_KEY (https://www.reed.co.uk/developers/)"
            "\nGreenhouse/Lever sources will still run without credentials.\n"
        )

    jobs = scrape_all()
    log.info(f"\nTotal relevant postings found: {len(jobs)}")

    if jobs:
        build_excel(jobs, OUTPUT_FILE)
    else:
        log.warning("No results — check credentials and try again.")

    log.info("Done.")
